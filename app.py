import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill
import csv
import io

# ======================
# CONFIG PAGE
# ======================
st.set_page_config(
    page_title="Transformation de facturation",
    layout="wide"
)

st.title("🧾 Outil de transformation de facturation")
st.caption("Importez vos fichiers puis cliquez sur Transformer.")

# ======================
# UPLOAD FICHIERS
# ======================
doc1 = st.file_uploader(
    "📄 Fichier de facturation (Doc 1)",
    type=["csv", "xlsx"]
)

doc2 = st.file_uploader(
    "📄 Liste des raisons sociales (Doc 2)",
    type=["csv", "xlsx"]
)

# ======================
# LECTURE ROBUSTE
# ======================
def read_file(file):
    if file.name.lower().endswith(".xlsx"):
        return pd.read_excel(file)

    raw = file.read()
    file.seek(0)

    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin1")

    sniffer = csv.Sniffer()
    dialect = sniffer.sniff(text[:5000], delimiters=";,|\t")
    sep = dialect.delimiter

    return pd.read_csv(
        io.StringIO(text),
        sep=sep,
        engine="python",
        on_bad_lines="skip"
    )

# ======================
# TRANSFORMATION
# ======================
if doc1 and doc2 and st.button("🚀 Transformer les fichiers"):
    with st.spinner("⏳ Transformation en cours…"):

        # ----------------------
        # Lecture
        # ----------------------
        df = read_file(doc1)
        rs_df = read_file(doc2)

        # ----------------------
        # Normalisation
        # ----------------------
        df.iloc[:, 1] = df.iloc[:, 1].astype(str).str.strip()   # Raison sociale
        rs_df.iloc[:, 0] = rs_df.iloc[:, 0].astype(str).str.strip()

        df.iloc[:, 9] = pd.to_numeric(df.iloc[:, 9], errors="coerce")

        # ----------------------
        # Suppression lignes invalides
        # ----------------------
        df = df[
            df.iloc[:, 4].notna() &   # Status
            df.iloc[:, 6].notna()     # Date d’opération
        ].copy()

        # ----------------------
        # Base ALL (pour résumé)
        # ----------------------
        base_all = df[
            (df.iloc[:, 4] != "NOT INJECTED") &
            (df.iloc[:, 9] > 0)
        ].copy()

        # ----------------------
        # Filtre Doc 2
        # ----------------------
        in_doc2_all = base_all.iloc[:, 1].isin(rs_df.iloc[:, 0])
        base_doc2 = base_all[in_doc2_all].copy()

        # ======================
        # SYNTHÈSE GLOBALE
        # ======================
        service_col = df.columns[3]

        is_sms_all = base_all[service_col] == "SMS"
        is_vocal_all = base_all[service_col] == "VOCAL"

        summary = pd.DataFrame({
            "Catégorie": [
                "SMS – Raisons sociales du doc 2",
                "SMS – Autres raisons sociales",
                "Vocal – Raisons sociales du doc 2",
                "Vocal – Autres raisons sociales"
            ],
            "Nombre de messages": [
                base_all[is_sms_all & in_doc2_all].iloc[:, 9].sum(),
                base_all[is_sms_all & ~in_doc2_all].iloc[:, 9].sum(),
                base_all[is_vocal_all & in_doc2_all].iloc[:, 9].sum(),
                base_all[is_vocal_all & ~in_doc2_all].iloc[:, 9].sum(),
            ]
        })

        summary_display = summary.copy()
        summary_display["Nombre de messages"] = summary_display["Nombre de messages"].apply(
            lambda x: f"{int(x):,}".replace(",", " ")
        )

        # ======================
        # FACTURATION DÉTAILLÉE
        # ======================
        group_cols = [df.columns[1], df.columns[2]]

        agg = {
            df.columns[0]: "first",  # Plateforme
            df.columns[1]: "first",  # Raison sociale
            df.columns[2]: "first",  # Numéro opération
            df.columns[3]: "first",  # Type
            df.columns[4]: "first",  # Status
            df.columns[5]: "first",  # Nom opération
            df.columns[6]: "first",  # Date
            df.columns[7]: "first",  # Validation
            df.columns[8]: "first",  # Pays
            df.columns[9]: "sum",    # Nombre messages
        }

        df_final = (
            base_doc2
            .groupby(group_cols, as_index=False)
            .agg(agg)
        )

        df_final.columns = [
            "Plateforme",
            "Raison sociale",
            "Numéro d’opération",
            "Type",
            "Status",
            "Nom de l’opération",
            "Date d’opération",
            "Validation",
            "Pays",
            "Nombre de messages envoyés"
        ]

    # ======================
    # FEEDBACK UTILISATEUR
    # ======================
    st.success("✅ Transformation terminée avec succès")

    st.info(
        f"✔ {len(df_final):,}".replace(",", " ") + " lignes générées\n"
        f"✔ {df_final['Raison sociale'].nunique():,}".replace(",", " ") + " raisons sociales\n"
        f"✔ {df_final.shape[1]} colonnes en sortie"
    )

    # ======================
    # AFFICHAGE
    # ======================
    st.subheader("🔎 Facturation détaillée")
    st.dataframe(df_final, width="stretch")

    st.subheader("📊 Résumé global SMS / Vocal")
    st.dataframe(summary_display, width="stretch")

    st.info("Le résumé est inclus dans la deuxième feuille de l’Excel.")

    # ======================
    # EXPORT EXCEL
    # ======================
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Facturation détaillée")
        summary.to_excel(writer, index=False, sheet_name="Synthèse globale")

        ws = writer.sheets["Facturation détaillée"]
        fill_a = PatternFill("solid", fgColor="EEEEEE")
        fill_b = PatternFill("solid", fgColor="FFFFFF")

        last_rs, toggle = None, False
        for row in range(2, ws.max_row + 1):
            rs = ws.cell(row=row, column=2).value
            if rs != last_rs:
                toggle = not toggle
                last_rs = rs
            fill = fill_a if toggle else fill_b
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    st.download_button(
        "⬇️ Télécharger l’Excel final",
        data=output.getvalue(),
        file_name="facturation_finale.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
